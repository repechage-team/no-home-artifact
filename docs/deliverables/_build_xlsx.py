from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

FONT = "Malgun Gothic"
HEAD_FILL = PatternFill("solid", fgColor="1F6F5B")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1D2433")
BASE_FONT = Font(name=FONT, size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="D9E1E8")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# 담당자 색
COLOR = {"이정헌": "4F81BD", "최민식": "9BBB59", "전효준": "F79646", "공통": "A6A6A6"}

wb = Workbook()

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = CENTER; cell.border = BORDER

def put_rows(ws, rows, start=2):
    for r, rowdata in enumerate(rows, start=start):
        for c, val in enumerate(rowdata, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BASE_FONT; cell.alignment = WRAP; cell.border = BORDER

# ===== 시트 1: 요구사항명세 =====
ws1 = wb.active; ws1.title = "요구사항명세"
ws1["A1"] = "NoHome 요구사항 명세 (기능)"; ws1["A1"].font = TITLE_FONT
h1 = ["구분", "ID", "기능명", "설명", "우선순위", "담당", "완료 기준"]
ws1.append([])  # row2 spacer becomes header row3
for c, v in enumerate(h1, 1): ws1.cell(row=3, column=c, value=v)
style_header(ws1, 3, len(h1))
req = [
 ["필수","F701","실거래가 수집","매매·전월세 실거래가 공공데이터 API 수집·DB 적재","M","최민식","매매·전월세 DB 적재, 중복 skip"],
 ["필수","F702","실거래가 검색","지역·아파트명·거래유형·가격 조건 목록 조회","M","최민식","조건에 맞는 결과 반환"],
 ["필수","F712","회원 등록","이메일·비밀번호·이름·전화 가입","M","전효준","가입, 중복 이메일 거부"],
 ["필수","F713","회원 조회","로그인 사용자 본인 정보 조회","M","전효준","본인만 조회"],
 ["필수","F714","회원 수정","본인 정보 수정","M","전효준","본인만 허용 필드 수정"],
 ["필수","F715","회원 삭제","본인 회원 물리 삭제","M","전효준","물리 삭제, 본인만"],
 ["필수","F716","로그인 관리","로그인/로그아웃","M","전효준","로그인·로그아웃, 접근 제어"],
 ["필수","F720","주택 정보 관리","실거래가 기반 주택 기본정보 조회","M","최민식","houses/house_deals 연계"],
 ["필수","F-MAP","검색결과 지도 표시","Kakao Map 마커 표시","M","전효준","geocoding 후 마커"],
 ["추가","F-AI1","AI 질문 모드","자연어 질문에 데이터 요약 답변","S","이정헌","통계 요약 응답"],
 ["추가","F-AI2","AI 에이전트 모드","자연어로 필터·검색·화면 조작","S","이정헌","6액션 실행"],
 ["추가","F-AI3","단일 어시스턴트 통합","단일 /assistant tool calling 분기","S","이정헌","모드토글 없이 자동 분기"],
 ["추가","F-AI4","단기 대화기억","세션 휘발 멀티턴(InMemory window 10)","S","이정헌","맥락 유지, 종료 시 초기화"],
 ["추가","F-RENT","전월세 검색","dealMode 검색·가격필터","S","최민식","전세=보증금, 월세=보증금+월세"],
 ["추가","F-LIVE","라이브 검색 성능","응답 우선+백그라운드 적재","S","최민식","미캐시 첫 조회 지연 완화"],
 ["추가","F-INT","관심 지역","회원 관심 지역 등록·관리","S","전효준","회원-지역 저장·조회"],
 ["추가","F-ADM","회원검색 관리자권한","/members/search 관리자 전용","S","전효준","일반회원 403"],
 ["추가","F-PWD","비밀번호 찾기","비밀번호 재설정","S","전효준","재설정 처리"],
]
put_rows(ws1, req, start=4)
# 담당 색
for r in range(4, 4+len(req)):
    owner = ws1.cell(row=r, column=6).value
    if owner in COLOR:
        ws1.cell(row=r, column=6).fill = PatternFill("solid", fgColor=COLOR[owner])
        ws1.cell(row=r, column=6).font = Font(name=FONT, size=10, color="FFFFFF", bold=True)
widths1 = [8,8,18,40,8,10,34]
for i,w in enumerate(widths1,1): ws1.column_dimensions[chr(64+i)].width = w

# ===== 시트 2: WBS =====
ws2 = wb.create_sheet("WBS")
ws2["A1"] = "NoHome WBS (일정 기준: git 실측 2026-06-12~06-24)"; ws2["A1"].font = TITLE_FONT
h2 = ["ID","작업명","담당","시작","종료","상태","참조"]
for c,v in enumerate(h2,1): ws2.cell(row=3, column=c, value=v)
style_header(ws2, 3, len(h2))
wbs = [
 ["1","프로젝트 기반·인프라","공통","2026-06-12","2026-06-14","완료","M0~M1, S0~1"],
 ["1.1","문서·하네스(PRD·spec·plan)","공통","2026-06-12","2026-06-12","완료","S0"],
 ["1.2","Spring Boot·MyBatis 골격, /api/health","공통","2026-06-12","2026-06-13","완료","S1"],
 ["1.3","MySQL·Docker compose, .env","최민식","2026-06-13","2026-06-14","완료","S1"],
 ["1.4","폴더구조(Backend/Frontend/Artifact)","최민식","2026-06-14","2026-06-14","완료","S10"],
 ["2","공공데이터·실거래가 검색","최민식","2026-06-14","2026-06-24","완료","M2, S2~5"],
 ["2.1","실거래가 스키마","최민식","2026-06-14","2026-06-14","완료","S2"],
 ["2.2","매매 적재·api_row_hash 중복제거","최민식","2026-06-22","2026-06-22","완료","S3, #13"],
 ["2.3","통합 검색 API·자동임포트","최민식","2026-06-22","2026-06-22","완료","S4~5"],
 ["2.4","전월세 검색(dealMode·스키마 확장)","최민식","2026-06-22","2026-06-23","완료","#23"],
 ["2.5","라이브 검색 파이프라인","최민식","2026-06-23","2026-06-23","완료","#27"],
 ["2.6","라이브 검색 성능개선","최민식","2026-06-24","2026-06-24","완료","#30"],
 ["3","회원·인증·지도","전효준","2026-06-18","2026-06-24","완료","M3·M4, S6~9"],
 ["3.1","토큰기반 회원 인증(가입/로그인/CRUD)","전효준","2026-06-18","2026-06-18","완료","S6, #1"],
 ["3.2","회원 검색 + 비밀번호 찾기","전효준","2026-06-23","2026-06-23","완료","#22"],
 ["3.3","공지사항 + 관심지역","전효준","2026-06-23","2026-06-23","완료","#26"],
 ["3.4","Kakao 지도·geocoding·마커","전효준","2026-06-23","2026-06-24","완료","S9, #33"],
 ["3.5","회원검색 관리자권한·지도안내","전효준","2026-06-24","2026-06-24","완료","#33"],
 ["4","AI 어시스턴트","이정헌","2026-06-21","2026-06-24","완료","추가기능 트랙"],
 ["4.1","Spring AI 챗봇·질문모드·사용량 제한","이정헌","2026-06-21","2026-06-22","완료","#2"],
 ["4.2","GMS graceful·JWT fail-closed·구보정","이정헌","2026-06-22","2026-06-22","완료","#5,6,7"],
 ["4.3","에이전트 모드(구조화 명령)","이정헌","2026-06-22","2026-06-23","완료","#16"],
 ["4.4","에이전트 Phase2(필터·액션 확장)","이정헌","2026-06-23","2026-06-23","완료","#18"],
 ["4.5","공공데이터 resultCode 000 보정","이정헌","2026-06-23","2026-06-23","완료","#20"],
 ["4.6","단기 대화기억(InMemory window 10)","이정헌","2026-06-23","2026-06-23","완료","#28"],
 ["4.7","Tool Calling 재설계(/assistant 단일화)","이정헌","2026-06-24","2026-06-24","완료","#29"],
 ["5","프론트엔드 화면·UX","공통","2026-06-14","2026-06-24","완료","M4, S7~12"],
 ["5.1","메인 검색/목록/상세 shell·지도 placeholder","최민식","2026-06-14","2026-06-22","완료","S7"],
 ["5.2","검색 필터 UI(전월세·슬라이더·정렬)","최민식","2026-06-22","2026-06-23","완료","S11"],
 ["5.3","회원 패널·관심지역 UI","전효준","2026-06-23","2026-06-24","완료","FE #8,10,16"],
 ["5.4","AI 챗 위젯·에이전트 런타임·리사이즈","이정헌","2026-06-22","2026-06-24","완료","FE #13,17"],
 ["5.5","법정동 select·결과카드·mojibake 보정","최민식","2026-06-23","2026-06-24","완료","S11~12"],
 ["6","문서·발표·제출","공통","2026-06-12","2026-06-24","진행중","M7 준비"],
 ["6.1","Sprint·AI 설계/구현/트러블슈팅 문서","이정헌","2026-06-12","2026-06-24","완료","docs/*"],
 ["6.2","발표자료(구성안·대본·시연)","공통","2026-06-24","2026-06-24","완료","presentation/"],
 ["6.3","요구사항 명세서·WBS·간트","공통","2026-06-24","2026-06-24","진행중","deliverables/"],
]
put_rows(ws2, wbs, start=4)
for r in range(4, 4+len(wbs)):
    idv = ws2.cell(row=r, column=1).value
    owner = ws2.cell(row=r, column=3).value
    # L2(정수 ID) 굵게 강조
    if "." not in idv:
        for c in range(1,8):
            ws2.cell(row=r, column=c).font = Font(name=FONT, size=10, bold=True)
        for c in range(1,8):
            ws2.cell(row=r, column=c).fill = PatternFill("solid", fgColor="EEF3F0")
    if owner in COLOR:
        oc = ws2.cell(row=r, column=3)
        oc.fill = PatternFill("solid", fgColor=COLOR[owner])
        oc.font = Font(name=FONT, size=10, color="FFFFFF", bold=True)
widths2 = [6,42,10,12,12,8,16]
for i,w in enumerate(widths2,1): ws2.column_dimensions[chr(64+i)].width = w

# ===== 시트 3: 간트 =====
ws3 = wb.create_sheet("간트")
ws3["A1"] = "NoHome 간트 (담당자 색 막대, 2026-06-12~06-24)"; ws3["A1"].font = TITLE_FONT
days = [date(2026,6,d) for d in range(12,25)]
# 헤더
ws3.cell(row=3, column=1, value="작업")
ws3.cell(row=3, column=2, value="담당")
for i,d in enumerate(days):
    ws3.cell(row=3, column=3+i, value=d.strftime("%m-%d"))
style_header(ws3, 3, 2+len(days))
gtasks = [
 ("문서·하네스","공통","2026-06-12","2026-06-12"),
 ("Boot·MyBatis 골격","공통","2026-06-12","2026-06-13"),
 ("MySQL·Docker·폴더","최민식","2026-06-13","2026-06-14"),
 ("실거래가 스키마","최민식","2026-06-14","2026-06-14"),
 ("매매 적재·중복제거","최민식","2026-06-22","2026-06-22"),
 ("검색 API·자동임포트","최민식","2026-06-22","2026-06-22"),
 ("전월세 검색","최민식","2026-06-22","2026-06-23"),
 ("라이브 검색 파이프라인","최민식","2026-06-23","2026-06-23"),
 ("라이브 검색 성능개선","최민식","2026-06-24","2026-06-24"),
 ("토큰기반 회원 인증","전효준","2026-06-18","2026-06-18"),
 ("회원검색·비밀번호","전효준","2026-06-23","2026-06-23"),
 ("공지·관심지역","전효준","2026-06-23","2026-06-23"),
 ("Kakao 지도·geocoding","전효준","2026-06-23","2026-06-24"),
 ("관리자권한·지도안내","전효준","2026-06-24","2026-06-24"),
 ("Spring AI 챗봇·질문모드","이정헌","2026-06-21","2026-06-22"),
 ("보안·graceful·구보정","이정헌","2026-06-22","2026-06-22"),
 ("에이전트 모드","이정헌","2026-06-22","2026-06-23"),
 ("에이전트 Phase2","이정헌","2026-06-23","2026-06-23"),
 ("단기 대화기억","이정헌","2026-06-23","2026-06-23"),
 ("Tool Calling 재설계","이정헌","2026-06-24","2026-06-24"),
 ("메인 검색 shell·필터 UI","최민식","2026-06-14","2026-06-22"),
 ("AI 챗 위젯·회원 UI","공통","2026-06-22","2026-06-24"),
 ("발표자료","공통","2026-06-24","2026-06-24"),
 ("명세서·WBS·간트","공통","2026-06-24","2026-06-24"),
]
for r,(name,owner,s,e) in enumerate(gtasks, start=4):
    ws3.cell(row=r, column=1, value=name).font = BASE_FONT
    ws3.cell(row=r, column=1).border = BORDER
    oc = ws3.cell(row=r, column=2, value=owner)
    oc.font = Font(name=FONT, size=9, color="FFFFFF", bold=True)
    oc.fill = PatternFill("solid", fgColor=COLOR[owner]); oc.alignment = CENTER; oc.border = BORDER
    sd = date.fromisoformat(s); ed = date.fromisoformat(e)
    for i,d in enumerate(days):
        cell = ws3.cell(row=r, column=3+i)
        cell.border = BORDER
        if sd <= d <= ed:
            cell.fill = PatternFill("solid", fgColor=COLOR[owner])
ws3.column_dimensions["A"].width = 26
ws3.column_dimensions["B"].width = 9
for i in range(len(days)):
    ws3.column_dimensions[chr(67+i)].width = 5.5
# 범례
lr = 4+len(gtasks)+1
ws3.cell(row=lr, column=1, value="범례").font = Font(name=FONT, bold=True, size=10)
for j,(k,v) in enumerate(COLOR.items()):
    c = ws3.cell(row=lr, column=2+j, value=k)
    c.fill = PatternFill("solid", fgColor=v); c.font = Font(name=FONT, size=9, color="FFFFFF", bold=True)
    c.alignment = CENTER

# ===== 시트 4: 담당·일정 =====
ws4 = wb.create_sheet("담당·일정")
ws4["A1"] = "담당자별 요약 (git 실측)"; ws4["A1"].font = TITLE_FONT
h4 = ["담당","영역","활동 기간(git)","커밋수(3repo)","주요 PR"]
for c,v in enumerate(h4,1): ws4.cell(row=3, column=c, value=v)
style_header(ws4, 3, len(h4))
team = [
 ["이정헌","AI 어시스턴트·보안·공공데이터 보정","2026-06-12 ~ 06-24",53,"#2,5,6,7,16,18,20,28,29"],
 ["최민식","공공데이터·실거래가 검색·인프라","2026-06-14, 06-22 ~ 06-24",18,"#13,23,27,30"],
 ["전효준","회원·인증·지도·관심지역","2026-06-18, 06-23 ~ 06-24",11,"#1,22,26,33"],
]
put_rows(ws4, team, start=4)
for r in range(4,4+len(team)):
    oc = ws4.cell(row=r, column=1)
    oc.fill = PatternFill("solid", fgColor=COLOR[oc.value]); oc.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
ws4.cell(row=8, column=1, value="합계").font = Font(name=FONT, bold=True, size=10)
ws4.cell(row=8, column=4, value=53+18+11).font = Font(name=FONT, bold=True, size=10)  # 82, 3repo 합산 근사
widths4 = [10,34,24,14,26]
for i,w in enumerate(widths4,1): ws4.column_dimensions[chr(64+i)].width = w
ws4.cell(row=10, column=1, value="※ 커밋수는 backend+frontend+artifact 3개 repo 합산(중복 계정 포함 근사). PR 번호는 backend 기준.").font = Font(name=FONT, size=9, italic=True, color="5A6678")

import os
out = os.path.join(os.path.dirname(__file__), "no-home-deliverables.xlsx")
wb.save(out)
print("SAVED", out)
